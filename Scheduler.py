from datetime import timedelta
from prettytable import PrettyTable
from Lecture import Lecture
import tkinter as tk
from tkinter import ttk
import json
import openpyxl
from openpyxl.styles import Alignment, Font

def time_to_minutes(t):
    """Convert HH:MM time string to minutes."""
    return int(t.split(":")[0]) * 60 + int(t.split(":")[1])


def minutes_to_time(m):
    """Convert minutes back to HH:MM format."""
    return f"{m // 60:02d}:{m % 60:02d}"


def load_config():
    """Load class and break configurations from JSON files."""
    with open("class.config.json") as f:
        class_config = json.load(f)
    with open("break.config.json") as f:
        break_config = json.load(f)
    return class_config, break_config


def parse_break_times(break_config):
    """Convert break times to minutes."""
    breaks = []
    for break_type in break_config["Breaks"]:
        for timing in break_config["Breaks"][break_type]["timeing"]:
            start_time, end_time = timing.split("-")
            breaks.append((time_to_minutes(start_time), time_to_minutes(end_time)))

    print(breaks)
    return breaks


def is_time_available(start_time, duration, day_schedule, breaks, class_end_time):
    """Check if the time slot is available, considering the breaks and existing schedule."""
    end_time = start_time + duration
    if end_time > class_end_time:
        return False
    for break_start, break_end in breaks:
        if not (end_time <= break_start or start_time >= break_end):
            return False
    for lecture in day_schedule:
        lec_start, lec_end = map(time_to_minutes, lecture["Time"].split("-"))
        if not (end_time <= lec_start or start_time >= lec_end):
            return False
    return True


def add_breaks_to_schedule(schedule, breaks):
    """Add break times to the schedule."""
    for day in schedule:
        for break_start, break_end in breaks:
            schedule[day].append(
                {
                    "Course Name": "Break",
                    "Time": f"{minutes_to_time(break_start)}-{minutes_to_time(break_end)}",
                    "Professor Name": "",
                    "Corp": "Break",
                }
            )


def create_schedule(lectures, class_config, break_config):
    """Generate a weekly schedule based on class and break configurations."""
    # Initialize schedule for each day
    schedule = {day: [] for day in class_config["DAYS"]}

    # Parse config values
    class_start_time = time_to_minutes(class_config["START_TIME"])
    class_end_time = time_to_minutes(class_config["END_TIME"])
    lecture_duration = class_config["LECTURE_DURATION"]
    max_lectures_per_day = class_config["MAX_LECTURES_PER_DAY"]
    max_same_lecture_count = class_config["MAX_SAME_LECTURE_COUNT_IN_SINGLE_DAY"]
    gap_between_lectures = break_config["GAP_TIME_BETWEEN_LECTURES"]
    breaks = parse_break_times(break_config)

    print(breaks)

    days = class_config["DAYS"]
    current_day_index = days.index(class_config["WEEK_START_DAY"])

    def add_lecture(day, course_name, professor_name, start_time, end_time):
        """Helper to add a lecture to the schedule for a specific day."""
        schedule[day].append(
            {
                "Course Name": course_name,
                "Time": f"{minutes_to_time(start_time)}-{minutes_to_time(end_time)}",
                "Professor Name": professor_name,
                "Corp": course_name,
            }
        )

    for lecture in lectures:
        course_name = lecture.course_name
        professor_name = lecture.professor
        total_lectures = lecture.hpw
        same_lecture_count = 0  # Track how many consecutive lectures of the same course

        while total_lectures > 0:
            day = days[current_day_index]
            current_time = class_start_time
            lectures_today = 0

            while lectures_today < max_lectures_per_day and total_lectures > 0:
                if same_lecture_count < max_same_lecture_count:
                    if is_time_available(
                        current_time,
                        lecture_duration,
                        schedule[day],
                        breaks,
                        class_end_time,
                    ):
                        end_time = current_time + lecture_duration
                        add_lecture(
                            day, course_name, professor_name, current_time, end_time
                        )
                        current_time = end_time + gap_between_lectures
                        lectures_today += 1
                        same_lecture_count += 1
                        total_lectures -= 1
                    else:
                        current_time += 5  # Skip 5 minutes if time slot not available
                else:
                    # Reset count after max same lectures and move to the next day
                    same_lecture_count = 0
                    current_day_index = (current_day_index + 1) % len(days)
                    break

                # Move to the next day if the current day is fully scheduled or time runs out
                if (
                    current_time >= class_end_time
                    or lectures_today >= max_lectures_per_day
                ):
                    current_day_index = (current_day_index + 1) % len(days)
                    current_time = class_start_time
                    same_lecture_count = 0

    # Add breaks to the schedule
    add_breaks_to_schedule(schedule, breaks)

    return schedule


def print_schedule(schedule):
    """Display the schedule in a readable format."""
    for day, lectures in schedule.items():
        lectures = sorted(
            lectures, key=lambda x: time_to_minutes(x["Time"].split("-")[0])
        )
        print(f"\n{day}")
        print("=" * 40)
        for lecture in lectures:
            print(f"Course Name: {lecture['Course Name']}")
            print(f"Time: {lecture['Time']}")
            print(f"Professor Name: {lecture['Professor Name']}")
            print(f"Corp: {lecture['Corp']}")
            print("-" * 40)


def print_schedule_table(schedule):
    """Display the schedule in a table format."""
    for day, lectures in schedule.items():
        lectures = sorted(
            lectures, key=lambda x: time_to_minutes(x["Time"].split("-")[0])
        )
        # Create a table for each day
        table = PrettyTable()
        table.field_names = ["Course Name", "Time", "Professor Name", "Corp"]
        table.title = day  # Set the day as the title of the table

        for lecture in lectures:
            table.add_row(
                [
                    lecture["Course Name"],
                    lecture["Time"],
                    lecture["Professor Name"],
                    lecture["Corp"],
                ]
            )

        # Print the table for the day
        print(table)
        print("\n")


def create_gui_schedule(schedule):
    """Create a GUI window to display the schedule in table format."""
    # Initialize the main window
    root = tk.Tk()
    root.title("Weekly Schedule")

    # Create a frame to hold the table
    frame = ttk.Frame(root, padding="10")
    frame.grid(row=0, column=0)

    # Define the headers
    headers = ["Day", "Course Name", "Timing", "Professor"]
    for col, header in enumerate(headers):
        label = ttk.Label(frame, text=header, font=("Arial", 12, "bold"))
        label.grid(row=0, column=col, padx=10, pady=5)

    # Fill in the schedule data for each day
    row_index = 1
    for day, lectures in schedule.items():
        day_label = ttk.Label(frame, text=day, font=("Arial", 12, "bold"))
        day_label.grid(row=row_index, column=0, padx=10, pady=5, sticky="w")

        # Sort lectures based on time
        lectures = sorted(
            lectures, key=lambda x: time_to_minutes(x["Time"].split("-")[0])
        )

        for lecture in lectures:
            course_label = ttk.Label(
                frame, text=lecture["Course Name"], font=("Arial", 10)
            )
            course_label.grid(row=row_index, column=1, padx=10, pady=5)

            timing_label = ttk.Label(frame, text=lecture["Time"], font=("Arial", 10))
            timing_label.grid(row=row_index, column=2, padx=10, pady=5)

            professor_label = ttk.Label(
                frame, text=lecture["Professor Name"], font=("Arial", 10)
            )
            professor_label.grid(row=row_index, column=3, padx=10, pady=5)

            row_index += 1

        row_index += 1  # Leave a gap after each day's schedule

    # Start the GUI loop
    root.mainloop()


def save_schedule_to_excel(schedule, filename="schedule.xlsx"):
    """Save the schedule to an Excel file."""
    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Weekly Schedule"

    # Set the headers
    headers = ["Day", "Course Name", "Time", "Professor Name", "Corp"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Fill in the schedule data
    row_num = 2  # Start filling data from the second row
    for day, lectures in schedule.items():
        lectures = sorted(
            lectures, key=lambda x: time_to_minutes(x["Time"].split("-")[0])
        )
        for lecture in lectures:
            sheet.cell(row=row_num, column=1, value=day)
            sheet.cell(row=row_num, column=2, value=lecture["Course Name"])
            sheet.cell(row=row_num, column=3, value=lecture["Time"])
            sheet.cell(row=row_num, column=4, value=lecture["Professor Name"])
            sheet.cell(row=row_num, column=5, value=lecture["Corp"])
            row_num += 1

    # Auto-adjust column widths
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column].width = adjusted_width

    # Save the workbook to the given filename
    wb.save(filename)
    print(f"Schedule saved to {filename} successfully!")


if __name__ == "__main__":
    class_config, break_config = load_config()
    lectures = [
        Lecture(
            name="Introduction to Programming",
            corp="Dr. Shaun Murphy",
            credit=3,
            professor="Dr. Clair Brown",
            hpw=5,
        ),
        Lecture(
            name="Data Structures",
            corp="Dr. Clair Brown",
            credit=4,
            professor="Dr. Kalu Jared",
            hpw=4,
        ),
        Lecture(
            name="Algorithms",
            corp="Dr. Neil Melendez",
            credit=4,
            professor="Dr. Shaun Murphy",
            hpw=3,
        ),
        Lecture(
            name="Database Systems",
            corp="Dr. Lim Audrey",
            credit=4,
            professor="Dr. Glassman Aaron",
            hpw=6,
        ),
        Lecture(
            name="Operating Systems",
            corp="Dr. Kalu Jared",
            credit=4,
            professor="Dr. Park Alex",
            hpw=4,
        ),
        Lecture(
            name="Computer Networks",
            corp="Dr. Andrews Marcus",
            credit=3,
            professor="Dr. Lim Audrey",
            hpw=5,
        ),
        Lecture(
            name="Software Engineering",
            corp="Dr. Glassman Aaron",
            credit=4,
            professor="Dr. Neil Melendez",
            hpw=2,
        ),
        Lecture(
            name="Artificial Intelligence",
            corp="Dr. Park Alex",
            credit=4,
            professor="Dr. Andrews Marcus",
            hpw=3,
        ),
    ]

    schedule = create_schedule(lectures, class_config, break_config)
    # print_schedule(schedule)
    # print_schedule_table(schedule)
    # create_gui_schedule(schedule)
    save_schedule_to_excel(schedule)
